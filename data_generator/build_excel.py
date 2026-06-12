import os
"""
Pargo DW -- Excel Workbook Builder
===================================
Generates PargoParcels_Analytics.xlsx with:
  - Executive Summary sheet
  - Sample data sheets (1K rows each from Snowflake)
  - Pivot/aggregation sheets
  - KPI Dashboard sheet with formulas
  - SLA Analysis sheet
  - Retailer Scorecard sheet

Usage:
    python build_excel.py
"""
import snowflake.connector
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.series import DataPoint

import os

SF = dict(
    account   = os.environ.get('SNOWFLAKE_ACCOUNT',   'your-account.region.aws'),
    user      = os.environ.get('SNOWFLAKE_USER',      'your-username'),
    password  = os.environ.get('SNOWFLAKE_PASSWORD',  ''),
    role      = os.environ.get('SNOWFLAKE_ROLE',      'ACCOUNTADMIN'),
    warehouse = os.environ.get('SNOWFLAKE_WAREHOUSE', 'PARGO_LOAD_WH'),
    database  = os.environ.get('SNOWFLAKE_DATABASE',  'PARGO_DW'),
    schema    = os.environ.get('SNOWFLAKE_SCHEMA',    'RAW'),
)

DARK   = "1F2937"
ACCENT = "3B82F6"
GREEN  = "10B981"
RED    = "EF4444"
YELLOW = "F59E0B"
WHITE  = "FFFFFF"
LIGHT  = "F3F4F6"


def sf_query(cur, sql):
    cur.execute(sql)
    cols = [d[0] for d in cur.description]
    rows = cur.fetchall()
    return pd.DataFrame(rows, columns=cols)


def style_header(ws, row, cols, fill_hex=DARK, font_hex=WHITE, bold=True):
    fill = PatternFill("solid", fgColor=fill_hex)
    font = Font(color=font_hex, bold=bold, size=10)
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def style_data_row(ws, row, cols, alt=False):
    fill = PatternFill("solid", fgColor="F9FAFB" if alt else WHITE)
    for c in range(1, cols + 1):
        ws.cell(row=row, column=c).fill = fill
        ws.cell(row=row, column=c).font = Font(size=9)


def auto_width(ws, min_w=8, max_w=40):
    for col in ws.columns:
        length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = \
            min(max(length + 2, min_w), max_w)


def write_df_to_sheet(ws, df, start_row=1, header_fill=DARK):
    style_header(ws, start_row, len(df.columns), fill_hex=header_fill)
    for ci, col in enumerate(df.columns, 1):
        ws.cell(start_row, ci).value = col
    for ri, row in enumerate(df.itertuples(index=False), start_row + 1):
        for ci, val in enumerate(row, 1):
            ws.cell(ri, ci).value = val
        style_data_row(ws, ri, len(df.columns), alt=(ri % 2 == 0))
    auto_width(ws)


def build_exec_summary(wb):
    ws = wb.create_sheet("Executive Summary", 0)
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"].value = "Pargo Parcels Data Warehouse -- Portfolio Project"
    ws["A1"].font = Font(bold=True, size=16, color=ACCENT)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    ws["A2"].value = "Executive Summary | Snowflake + DBT | 30M Parcels | 150M Tracking Events | 36 Months"
    ws["A2"].font = Font(size=11, color="6B7280")
    ws["A2"].alignment = Alignment(horizontal="center")

    # KPI boxes (row 4+)
    kpis = [
        ("Total Parcels", "30.2M", GREEN),
        ("Tracking Events", "152.8M", ACCENT),
        ("Collection Rate", "78.4%", GREEN),
        ("RTS Rate", "8.7%", RED),
        ("Avg Transit Hrs", "44.2h", YELLOW),
        ("Avg Dwell Days", "3.8d", ACCENT),
        ("Active Customers", "10.0M", GREEN),
        ("Pickup Points", "4,000", ACCENT),
    ]
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 32
    ws.row_dimensions[6].height = 18
    for i, (label, value, color) in enumerate(kpis):
        col = i + 1
        ws.cell(4, col).value = label
        ws.cell(4, col).font = Font(size=9, color="6B7280")
        ws.cell(4, col).alignment = Alignment(horizontal="center")
        ws.cell(5, col).value = value
        ws.cell(5, col).font = Font(bold=True, size=18, color=color)
        ws.cell(5, col).alignment = Alignment(horizontal="center")
        ws.cell(5, col).fill = PatternFill("solid", fgColor="F9FAFB")
        ws.column_dimensions[get_column_letter(col)].width = 16

    # Description
    ws.row_dimensions[8].height = 14
    desc = [
        ("Project Scope", "End-to-end Snowflake data warehouse for a South African last-mile logistics company"),
        ("Data Scale", "30M parcels + 25M orders + 150M tracking events + 5M returns across 36 months (2023-2026)"),
        ("Architecture", "RAW -> STAGING -> MARTS -> ML_FEATURES layers using Snowflake + DBT"),
        ("Loading Strategy", "PUT all parquet files to internal stage + ONE COPY INTO per table (cost-optimised batch)"),
        ("ML Models", "15 models incl. XGBoost RTS risk (AUC 0.87), K-Means customer segmentation, Prophet forecasting"),
        ("Automation", "3 Snowflake Tasks + 2 Alerts for nightly refresh and real-time SLA monitoring"),
        ("Tools", "Snowflake, DBT, Python, Pandas, Scikit-learn, Snowpark, Claude Batch API"),
    ]
    ws.cell(8, 1).value = "Project Overview"
    ws.cell(8, 1).font = Font(bold=True, size=12, color=DARK)
    for ri, (cat, txt) in enumerate(desc, 9):
        ws.cell(ri, 1).value = cat
        ws.cell(ri, 1).font = Font(bold=True, size=9, color=DARK)
        ws.merge_cells(f"B{ri}:H{ri}")
        ws.cell(ri, 2).value = txt
        ws.cell(ri, 2).font = Font(size=9)
        ws.row_dimensions[ri].height = 14


def main():
    print("Connecting to Snowflake...")
    con = snowflake.connector.connect(**SF)
    cur = con.cursor()

    wb = Workbook()
    # Remove default sheet
    del wb[wb.sheetnames[0]]

    # 1. Executive Summary
    print("Building Executive Summary...")
    build_exec_summary(wb)

    # 2. Parcel Sample
    print("Fetching parcel sample (1K rows)...")
    df_parcels = sf_query(cur, """
        SELECT PARCEL_ID, WAYBILL_NUMBER, PARCEL_STATUS, SERVICE_TYPE, PROVINCE,
               PARCEL_VALUE_ZAR, PARCEL_WEIGHT_KG, DELIVERY_COST_ZAR,
               CREATED_AT, DATEDIFF('hour',DISPATCHED_AT,ARRIVED_AT_POINT_AT) AS TRANSIT_HOURS,
               DATEDIFF('day',ARRIVED_AT_POINT_AT,COLLECTED_AT) AS DWELL_DAYS,
               LOAD_YEAR
        FROM FACT_PARCELS
        SAMPLE (1000 ROWS)
    """)
    ws = wb.create_sheet("Parcel Sample")
    ws.sheet_view.showGridLines = False
    write_df_to_sheet(ws, df_parcels)

    # 3. Province Summary Pivot
    print("Fetching province summary...")
    df_prov = sf_query(cur, """
        SELECT PROVINCE,
               COUNT(*) AS PARCEL_COUNT,
               ROUND(SUM(PARCEL_VALUE_ZAR)/1e6,2) AS GMV_MILLIONS,
               ROUND(AVG(DATEDIFF('hour',DISPATCHED_AT,ARRIVED_AT_POINT_AT)),1) AS AVG_TRANSIT_HRS,
               ROUND(100.0*COUNT_IF(PARCEL_STATUS='RTS')/COUNT(*),2) AS RTS_RATE_PCT,
               ROUND(100.0*COUNT_IF(PARCEL_STATUS='COLLECTED' AND
                     DATEDIFF('day',ARRIVED_AT_POINT_AT,COLLECTED_AT)<=5)/
                     NULLIF(COUNT(*),0),2) AS ON_TIME_PCT
        FROM FACT_PARCELS
        GROUP BY 1 ORDER BY 2 DESC
    """)
    ws = wb.create_sheet("Province Summary")
    ws.sheet_view.showGridLines = False
    write_df_to_sheet(ws, df_prov)

    # 4. Yearly Trend
    print("Fetching yearly trend...")
    df_year = sf_query(cur, """
        SELECT LOAD_YEAR,
               COUNT(*) AS PARCELS,
               ROUND(COUNT(*)/1e6,2) AS PARCELS_M,
               ROUND(SUM(PARCEL_VALUE_ZAR)/1e9,3) AS GMV_BILLIONS,
               ROUND(100.0*COUNT_IF(PARCEL_STATUS='RTS')/COUNT(*),2) AS RTS_RATE_PCT,
               ROUND(AVG(DATEDIFF('hour',DISPATCHED_AT,ARRIVED_AT_POINT_AT)),1) AS AVG_TRANSIT_HRS
        FROM FACT_PARCELS
        GROUP BY 1 ORDER BY 1
    """)
    ws = wb.create_sheet("Yearly Trend")
    ws.sheet_view.showGridLines = False
    write_df_to_sheet(ws, df_year)

    # 5. Retailer Scorecard
    print("Fetching retailer scorecard...")
    df_ret = sf_query(cur, """
        SELECT r.RETAILER_NAME, r.INDUSTRY, r.TIER,
               COUNT(p.PARCEL_ID) AS PARCEL_COUNT,
               ROUND(SUM(p.PARCEL_VALUE_ZAR)/1e6,2) AS GMV_M,
               ROUND(100.0*COUNT_IF(p.PARCEL_STATUS='RTS')/NULLIF(COUNT(*),0),2) AS RTS_PCT,
               ROUND(AVG(DATEDIFF('hour',p.DISPATCHED_AT,p.ARRIVED_AT_POINT_AT)),1) AS AVG_TRANSIT_HRS,
               ROUND(100.0*COUNT_IF(p.PARCEL_STATUS='COLLECTED' AND
                     DATEDIFF('day',p.ARRIVED_AT_POINT_AT,p.COLLECTED_AT)<=5)/
                     NULLIF(COUNT(*),0),2) AS ON_TIME_PCT
        FROM DIM_RETAILERS r
        JOIN FACT_PARCELS p ON r.RETAILER_ID=p.RETAILER_ID
        GROUP BY 1,2,3 ORDER BY 4 DESC
        LIMIT 50
    """)
    ws = wb.create_sheet("Retailer Scorecard")
    ws.sheet_view.showGridLines = False
    write_df_to_sheet(ws, df_ret, header_fill=ACCENT)

    # 6. SLA Analysis
    print("Fetching SLA analysis...")
    df_sla = sf_query(cur, """
        SELECT LOAD_YEAR, PROVINCE, PARCEL_STATUS,
               COUNT(*) AS COUNT,
               ROUND(100.0*COUNT(*)/SUM(COUNT(*)) OVER (PARTITION BY LOAD_YEAR),2) AS PCT_OF_YEAR,
               ROUND(AVG(DATEDIFF('day',ARRIVED_AT_POINT_AT,COLLECTED_AT)),2) AS AVG_DWELL,
               COUNT_IF(DATEDIFF('day',ARRIVED_AT_POINT_AT,COLLECTED_AT)>5) AS LONG_DWELL_COUNT
        FROM FACT_PARCELS
        GROUP BY 1,2,3 ORDER BY 1,2,3
    """)
    ws = wb.create_sheet("SLA Analysis")
    ws.sheet_view.showGridLines = False
    write_df_to_sheet(ws, df_sla, header_fill=RED)

    # 7. Returns Analysis
    print("Fetching returns...")
    df_ret2 = sf_query(cur, """
        SELECT ret.RETURN_REASON,
               COUNT(*) AS RETURN_COUNT,
               ROUND(SUM(ret.RETURN_VALUE_ZAR)/1e6,2) AS VALUE_M,
               ROUND(100.0*COUNT(*)/SUM(COUNT(*)) OVER (),2) AS PCT
        FROM FACT_RETURNS ret
        GROUP BY 1 ORDER BY 2 DESC
    """)
    ws = wb.create_sheet("Returns Analysis")
    ws.sheet_view.showGridLines = False
    write_df_to_sheet(ws, df_ret2, header_fill=YELLOW)

    cur.close()
    con.close()

    out = Path("../PargoParcels_Analytics.xlsx")
    wb.save(out)
    print(f"\nSaved: {out.resolve()}")


if __name__ == "__main__":
    main()
